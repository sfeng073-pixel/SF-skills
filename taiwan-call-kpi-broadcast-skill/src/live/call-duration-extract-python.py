from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TAIWAN_GROUP = "台湾CC02组"
REQUIRED_HEADERS = {
    "小组",
    "CC",
    "通时目标",
    "通时",
    "通时达标率",
    "通次目标",
    "通次",
    "通次达标率",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_number(value: Any) -> float | None:
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_ratio(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return None
    try:
        return float(text)
    except ValueError:
        return None


def find_detail_header(sheet) -> tuple[int, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        headers = [clean_text(value) for value in row]
        if REQUIRED_HEADERS.issubset(set(headers)):
            return row_index, headers
        if row_index >= 40:
            break
    raise ValueError(f"Cannot find call-duration detail header with columns {sorted(REQUIRED_HEADERS)}")


def extract_rows(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header_row_index, headers = find_detail_header(sheet)
    rows: list[dict[str, Any]] = []
    current_group = ""

    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        padded = values + [None] * max(0, len(headers) - len(values))
        record = {headers[index]: padded[index] for index in range(len(headers)) if headers[index]}
        group = clean_text(record.get("小组"))
        if group:
            current_group = group
        elif current_group:
            record["小组"] = current_group

        group = clean_text(record.get("小组"))
        sales = clean_text(record.get("CC"))
        if group != TAIWAN_GROUP:
            continue
        if not sales or sales in {"汇总", "CC"}:
            continue

        rows.append(
            {
                "group": group,
                "sales": sales,
                "durationTarget": parse_number(record.get("通时目标")),
                "duration": parse_number(record.get("通时")),
                "durationRate": parse_ratio(record.get("通时达标率")),
                "callCountTarget": parse_number(record.get("通次目标")),
                "callCount": parse_number(record.get("通次")),
                "callCountRate": parse_ratio(record.get("通次达标率")),
            }
        )

    return rows


def main() -> int:
    if len(sys.argv) < 2:
        raise SystemExit("usage: call-duration-extract-python.py WORKBOOK_PATH")

    workbook_path = Path(sys.argv[1])
    rows = extract_rows(workbook_path)
    result = {
        "source_path": str(workbook_path),
        "group": TAIWAN_GROUP,
        "rows": rows,
        "summary": {
            "salesCount": len(rows),
        },
    }
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
