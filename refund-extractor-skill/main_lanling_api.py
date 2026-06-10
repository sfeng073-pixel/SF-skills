#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
蓝凌OA退费审批自动化工具 (纯HTTP API版本)
混合方案：蓝凌OA API获取列表 + 钉钉API获取详情/执行审批

作者: SOLO
"""

import os
import sys
import time
import json
import logging
import argparse
import re
import html
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============ 配置 ============
# 蓝凌OA域名（通过环境变量或直接修改）
LANLING_BASE = os.environ.get("LANLING_BASE", "https://your-lanling-domain.com")
LIST_API = f"{LANLING_BASE}/km/review/km_review_index/kmReviewIndex.do"

# 钉钉配置（通过环境变量设置，勿硬编码到代码中）
DINGTALK_APP_KEY = os.environ.get("DINGTALK_APP_KEY", "")
DINGTALK_APP_SECRET = os.environ.get("DINGTALK_APP_SECRET", "")
DINGTALK_PROCESS_CODE = os.environ.get("DINGTALK_PROCESS_CODE", "")

# 输出目录
OUTPUT_DIR = Path.home() / "refund-approval-automation" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Cookie文件（保存登录状态）
COOKIE_FILE = Path.home() / "refund-approval-automation" / ".lanling_cookies.json"

# 日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


# ============ Cookie管理 ============
class CookieManager:
    """管理蓝凌OA的Cookie登录状态"""

    def __init__(self):
        self.cookies = {}
        self.load()

    def load(self):
        """从文件加载Cookie"""
        if COOKIE_FILE.exists():
            try:
                with open(COOKIE_FILE, 'r') as f:
                    self.cookies = json.load(f)
                logger.info(f"已加载Cookie: {len(self.cookies)} 个")
            except Exception as e:
                logger.warning(f"加载Cookie失败: {e}")
                self.cookies = {}

    def save(self, cookie_dict: dict):
        """保存Cookie到文件"""
        self.cookies = cookie_dict
        try:
            with open(COOKIE_FILE, 'w') as f:
                json.dump(cookie_dict, f)
            logger.info("Cookie已保存")
        except Exception as e:
            logger.warning(f"保存Cookie失败: {e}")

    def get_dict(self):
        return self.cookies

    def get_header_string(self):
        """转为Cookie请求头格式"""
        return '; '.join([f"{k}={v}" for k, v in self.cookies.items()])


# ============ Excel处理 ============
class RefundExcelHandler:
    """退费审批Excel处理器"""

    DEFAULT_FIELDS = [
        "申请单编号", "申请人姓名", "申请人部门", "申请人工号",
        "申请日期", "学员大用户ID", "学员姓名", "退费学科",
        "豌豆学员ID", "退费订单ID", "退费订单名称",
        "合同付款金额", "合同金额", "退费金额", "最终可退金额",
        "退费原因（一级）", "退费原因（二级）", "退费说明",
        "换单结果", "关联工单", "退费附件文件",
        "直接主管", "直接主管意见", "间接主管", "间接主管意见",
        "审批完成时间"
    ]

    def __init__(self):
        self.output_dir = OUTPUT_DIR

    def get_path(self, date_str: str = None) -> Path:
        if date_str is None:
            date_str = datetime.now().strftime("%Y%m%d")
        return self.output_dir / f"refund_{date_str}.xlsx"

    def create_workbook(self, file_path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "退费审批明细"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, field in enumerate(self.DEFAULT_FIELDS, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        col_widths = [18, 10, 20, 12, 12, 14, 8, 10, 12, 18, 30,
                      12, 12, 12, 12, 12, 12, 30, 10, 15, 20,
                      12, 25, 12, 25, 15]
        for i, width in enumerate(col_widths, 1):
            col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
            ws.column_dimensions[col_letter].width = width

        wb.save(file_path)
        return wb

    def get_or_create(self, date_str: str = None):
        file_path = self.get_path(date_str)
        if not file_path.exists():
            return self.create_workbook(file_path)
        return load_workbook(file_path)

    def append(self, data: dict, date_str: str = None):
        file_path = self.get_path(date_str)
        wb = self.get_or_create(date_str)
        ws = wb.active
        next_row = ws.max_row + 1

        data_font = Font(size=10)
        data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, field in enumerate(self.DEFAULT_FIELDS, 1):
            value = data.get(field, "")
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        wb.save(file_path)
        logger.info(f"  数据已保存: {file_path}")
        return file_path


# ============ 蓝凌OA API封装 ============
class LanlingAPI:
    """蓝凌OA HTTP API封装"""

    def __init__(self):
        self.cookie_mgr = CookieManager()
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'application/json, text/javascript, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'X-Requested-With': 'XMLHttpRequest'
        })
        # 加载Cookie
        cookies = self.cookie_mgr.get_dict()
        if cookies:
            self.session.cookies.update(cookies)

    def _check_login(self) -> bool:
        """检查是否已登录"""
        try:
            resp = self.session.get(
                f"{LANLING_BASE}/km/review/",
                timeout=10,
                allow_redirects=False
            )
            # 如果返回200且没有跳转到登录页，说明已登录
            if resp.status_code == 200 and 'login' not in resp.url.lower():
                return True
            return False
        except Exception as e:
            logger.warning(f"检查登录状态失败: {e}")
            return False

    def ensure_login(self):
        """确保已登录，如果未登录则提示用户"""
        if self._check_login():
            logger.info("Cookie有效，已登录")
            return True

        logger.warning("Cookie已过期或无效，需要重新获取")
        print("\n" + "="*60)
        print("需要登录蓝凌OA")
        print("="*60)
        print("\n请按以下步骤操作：")
        print("1. 在Chrome中打开蓝凌OA页面（从钉钉工作台进入）")
        print("2. 按 F12 打开开发者工具")
        print("3. 切换到 Application(应用) 标签")
        print("4. 左侧找到 Storage -> Cookies -> 你的蓝凌OA域名")
        print("5. 复制所有Cookie的 Name 和 Value")
        print("6. 粘贴到下面（格式: name=value; name2=value2）")
        print("\n或者更简单的方法：")
        print("在Console中执行: document.cookie")
        print("把输出的字符串粘贴到下面")
        print("="*60 + "\n")

        cookie_str = input("请输入Cookie字符串: ").strip()
        if not cookie_str:
            logger.error("未输入Cookie，无法继续")
            return False

        # 解析Cookie
        cookie_dict = {}
        for item in cookie_str.split(';'):
            item = item.strip()
            if '=' in item:
                k, v = item.split('=', 1)
                cookie_dict[k.strip()] = v.strip()

        self.session.cookies.update(cookie_dict)
        self.cookie_mgr.save(cookie_dict)

        # 验证
        if self._check_login():
            logger.info("登录成功！")
            return True
        else:
            logger.error("登录验证失败，请检查Cookie")
            return False

    def get_list(self, mydoc: str = "all", page: int = 1, page_size: int = 15,
                 start_time: str = None, end_time: str = None) -> list:
        """获取审批列表

        Args:
            mydoc: 'all'=全部, 'approval'=待我审的
            page: 页码
            page_size: 每页条数
            start_time: 开始时间，格式 YYYY-MM-DD
            end_time: 结束时间，格式 YYYY-MM-DD
        """
        # 基础参数
        params = [
            ('method', 'list'),
            ('pagingSetting', ''),
            ('q.mydoc', mydoc),
            ('q.j_path', '/listAll'),
            ('orderby', 'docCreateTime'),
            ('ordertype', 'down'),
            ('pageno', str(page)),
            ('rowsize', str(page_size)),
            ('__seq', str(int(time.time() * 1000))),
            ('s_ajax', 'true')
        ]

        # 时间过滤：使用两个同名参数 q.docCreateTime（网页端实际抓包确认）
        if start_time and end_time:
            params.append(('q.docCreateTime', start_time))
            params.append(('q.docCreateTime', end_time))

        try:
            resp = self.session.get(LIST_API, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()

            # 解析数据
            datas = data.get('datas', [])
            items = []

            for row in datas:
                item = {}
                for cell in row:
                    col_name = cell.get('col', '')
                    value = cell.get('value', '')

                    # 清理HTML标签
                    if '<' in value:
                        value = re.sub(r'<[^>]+>', '', value)
                        value = html.unescape(value)
                        value = value.strip()

                    item[col_name] = value

                items.append(item)

            logger.info(f"获取到 {len(items)} 条记录")
            return items

        except Exception as e:
            logger.error(f"获取列表失败: {e}")
            return []

    def get_detail_page(self, fd_id: str) -> str:
        """获取详情页HTML"""
        url = f"{LANLING_BASE}/km/review/km_review_main/kmReviewMain.do?method=view&fdId={fd_id}"
        try:
            resp = self.session.get(url, timeout=30)
            resp.raise_for_status()
            return resp.text
        except Exception as e:
            logger.error(f"获取详情失败: {e}")
            return ""

    def get_attachments(self, fd_id: str) -> list:
        """获取附件列表"""
        url = f"{LANLING_BASE}/sys/attachment/sys_att_main/sysAttMain.do"
        params = {
            'method': 'list',
            'fdModelName': 'com.landray.kmss.km.review.model.KmReviewMain',
            'fdModelId': fd_id,
            's_ajax': 'true'
        }
        try:
            resp = self.session.get(url, params=params, timeout=30)
            data = resp.json()
            attachments = []
            for row in data.get('datas', []):
                att = {}
                for cell in row:
                    col = cell.get('col', '')
                    value = cell.get('value', '')
                    if col == 'fdFileName':
                        # 清理HTML标签
                        if '<' in value:
                            value = re.sub(r'<[^>]+>', '', value)
                        att['name'] = value.strip()
                    elif col == 'fdSize':
                        att['size'] = value
                if att.get('name'):
                    attachments.append(att)
            return attachments
        except Exception as e:
            logger.warning(f"获取附件列表失败: {e}")
            return []

    def extract_detail_from_html(self, html_content: str, fd_id: str = "") -> dict:
        """从详情页HTML中提取数据"""
        data = {field: "" for field in RefundExcelHandler.DEFAULT_FIELDS}

        if not html_content:
            return data

        # 提取标题中的信息
        title_match = re.search(r'<title>(.*?)</title>', html_content)
        if title_match:
            title = title_match.group(1)
            # 格式: 白沧海-【豌豆】退费申请-学员ID:19876637-20260602-263
            if '学员ID:' in title:
                parts = title.split('-')
                if len(parts) >= 2:
                    data['申请单编号'] = parts[-1] if parts[-1].isdigit() else ''

        # 提取表单表格数据
        # 查找所有表格行
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html_content, re.DOTALL)

        for row in rows:
            # 提取单元格
            cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL)
            if len(cells) >= 2:
                field_map = {
                    '申请人': '申请人姓名',
                    '申请人姓名': '申请人姓名',
                    '部门': '申请人部门',
                    '申请人部门': '申请人部门',
                    '工号': '申请人工号',
                    '申请人工号': '申请人工号',
                    '申请日期': '申请日期',
                    '学员大用户ID': '学员大用户ID',
                    '学员姓名': '学员姓名',
                    '退费学科': '退费学科',
                    '豌豆学员ID': '豌豆学员ID',
                    '退费订单ID': '退费订单ID',
                    '退费订单名称': '退费订单名称',
                    '合同付款金额': '合同付款金额',
                    '合同金额': '合同金额',
                    '退费金额': '退费金额',
                    '最终可退金额': '最终可退金额',
                    '退费原因（一级）': '退费原因（一级）',
                    '退费原因（二级）': '退费原因（二级）',
                    '退费说明': '退费说明',
                    '挽单结果': '换单结果',
                    '换单结果': '换单结果',
                    '关联工单': '关联工单',
                    '退费附件文件': '退费附件文件',
                }

                # 遍历所有单元格对（label-value对）
                # 表格可能是: [label1, value1, label2, value2] 或 [label1, value1]
                i = 0
                while i < len(cells) - 1:
                    label = re.sub(r'<[^>]+>', '', cells[i]).strip()
                    label = html.unescape(label)
                    value = re.sub(r'<[^>]+>', '', cells[i + 1]).strip()
                    value = html.unescape(value)

                    if label in field_map:
                        data[field_map[label]] = value
                        i += 2
                    else:
                        i += 1

        # 提取审批记录（流程处理）
        # 查找包含"流程处理"或"审批记录"的部分
        process_section = re.search(r'流程处理.*?(?=<div class="lui_tabpanel_nav")', html_content, re.DOTALL)
        if process_section:
            process_html = process_section.group(0)
            # 提取审批记录行
            process_rows = re.findall(r'<tr[^>]*>(.*?)</tr>', process_html, re.DOTALL)
            for p_row in process_rows:
                p_cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', p_row, re.DOTALL)
                if len(p_cells) >= 5:
                    node_name = re.sub(r'<[^>]+>', '', p_cells[1]).strip()
                    operator = re.sub(r'<[^>]+>', '', p_cells[2]).strip()
                    opinion = re.sub(r'<[^>]+>', '', p_cells[4]).strip()

                    if '直接主管' in node_name:
                        data['直接主管'] = operator
                        data['直接主管意见'] = opinion
                    elif '间接主管' in node_name:
                        data['间接主管'] = operator
                        data['间接主管意见'] = opinion

        # 获取附件列表
        if fd_id:
            attachments = self.get_attachments(fd_id)
            if attachments:
                data['退费附件文件'] = ', '.join([a['name'] for a in attachments])

        data['审批完成时间'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return data


# ============ 钉钉API封装 ============
class DingTalkAPI:
    """钉钉审批API封装"""

    def __init__(self):
        self.app_key = DINGTALK_APP_KEY
        self.app_secret = DINGTALK_APP_SECRET
        self.process_code = DINGTALK_PROCESS_CODE
        self.access_token = None

    def get_token(self):
        """获取access_token"""
        url = "https://oapi.dingtalk.com/gettoken"
        params = {
            "appkey": self.app_key,
            "appsecret": self.app_secret
        }
        resp = requests.get(url, params=params, timeout=30)
        data = resp.json()
        if data.get("errcode") == 0:
            self.access_token = data["access_token"]
            return self.access_token
        raise Exception(f"获取token失败: {data}")

    def get_instance_detail(self, instance_id: str) -> dict:
        """获取审批实例详情"""
        if not self.access_token:
            self.get_token()

        url = "https://oapi.dingtalk.com/topapi/processinstance/get"
        payload = {"process_instance_id": instance_id}
        headers = {"Content-Type": "application/json"}

        resp = requests.post(
            f"{url}?access_token={self.access_token}",
            json=payload, headers=headers, timeout=30
        )
        data = resp.json()

        if data.get("errcode") == 0:
            return data.get("process_instance", {})
        return {}

    def approve(self, instance_id: str, task_id: str = None) -> bool:
        """执行同意审批"""
        if not self.access_token:
            self.get_token()

        # 如果没有task_id，先获取详情
        if not task_id:
            detail = self.get_instance_detail(instance_id)
            tasks = detail.get("tasks", [])
            for task in tasks:
                if task.get("task_status") == "RUNNING":
                    task_id = task.get("task_id")
                    break

        if not task_id:
            logger.warning("未找到待办任务")
            return False

        url = "https://oapi.dingtalk.com/topapi/process/instance/execute"
        payload = {
            "request": {
                "process_instance_id": instance_id,
                "task_id": task_id,
                "result": "agree",
                "remark": "同意退费申请"
            }
        }
        headers = {"Content-Type": "application/json"}

        resp = requests.post(
            f"{url}?access_token={self.access_token}",
            json=payload, headers=headers, timeout=30
        )
        data = resp.json()

        if data.get("errcode") == 0:
            logger.info("  ✓ 钉钉API审批同意成功")
            return True
        else:
            logger.warning(f"  ✗ 钉钉API审批失败: {data.get('errmsg')}")
            return False


# ============ 主程序 ============
class RefundApprovalApp:
    """退费审批主应用"""

    def __init__(self):
        self.lanling = LanlingAPI()
        self.dingtalk = DingTalkAPI()
        self.excel = RefundExcelHandler()

    def run_daily(self):
        """每日审批任务"""
        logger.info(f"\n{'='*60}")
        logger.info(f"退费审批任务 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"{'='*60}\n")

        # 确保登录
        if not self.lanling.ensure_login():
            logger.error("登录失败，任务终止")
            return

        # 获取待我审的列表
        logger.info("获取'待我审的'列表...")
        items = self.lanling.get_list(mydoc="approval")

        if not items:
            logger.info("✓ 没有待审批的流程")
            return

        # 筛选【豌豆】退费申请
        refund_items = [
            item for item in items
            if item.get('fdTemplate.fdName', '') == '【豌豆】退费申请'
        ]

        logger.info(f"找到 {len(refund_items)} 个【豌豆】退费申请")

        if not refund_items:
            logger.info("✓ 没有待审批的【豌豆】退费申请")
            return

        # 处理每个退费申请
        approved_count = 0
        for item in refund_items:
            fd_id = item.get('fdId', '')
            subject = item.get('docSubject', '')
            fd_number = item.get('fdNumber', '')
            handler = item.get('handlerName', '')

            logger.info(f"\n处理: {subject}")
            logger.info(f"  申请单编号: {fd_number}")
            logger.info(f"  当前处理人: {handler}")

            # 获取详情页HTML
            html_content = self.lanling.get_detail_page(fd_id)

            # 提取数据
            data = self.lanling.extract_detail_from_html(html_content, fd_id)
            data['申请单编号'] = fd_number or data.get('申请单编号', '')

            # 保存到Excel
            self.excel.append(data)

            # 尝试用钉钉API审批
            # 注意：需要知道对应的钉钉instance_id，这里蓝凌OA的fd_id和钉钉的instance_id可能不同
            # 暂时跳过钉钉API审批，后续可以匹配
            logger.info("  数据已提取，跳过自动审批（需匹配钉钉instance_id）")

        logger.info(f"\n{'='*60}")
        logger.info(f"任务完成: 处理了 {len(refund_items)} 条【豌豆】退费申请")
        logger.info(f"{'='*60}\n")

    def run_extract_all(self, days: int = 30):
        """提取所有【豌豆】退费申请数据（用于复盘）

        Args:
            days: 提取最近N天的数据
        """
        logger.info(f"\n{'='*60}")
        logger.info(f"退费数据提取任务 - 最近{days}天")
        logger.info(f"{'='*60}\n")

        if not self.lanling.ensure_login():
            logger.error("登录失败")
            return

        # 计算时间范围
        end_dt = datetime.now()
        start_dt = end_dt - timedelta(days=days)
        start_time = start_dt.strftime('%Y-%m-%d')
        end_time = end_dt.strftime('%Y-%m-%d')
        logger.info(f"时间范围: {start_time} ~ {end_time}")

        # 获取全部列表（API层面时间过滤）
        logger.info("获取全部流程列表（API时间过滤）...")
        all_items = []
        seen_ids = set()
        page = 1

        while True:
            items = self.lanling.get_list(
                mydoc="all", page=page, page_size=15,
                start_time=start_time, end_time=end_time
            )
            if not items:
                break

            new_count = 0
            for item in items:
                fd_id = item.get('fdId', '')
                if fd_id in seen_ids:
                    continue
                seen_ids.add(fd_id)

                # 筛选【豌豆】退费申请
                if item.get('fdTemplate.fdName', '') != '【豌豆】退费申请':
                    continue

                all_items.append(item)
                new_count += 1

            logger.info(f"  第{page}页: 本页{len(items)}条，新增{new_count}条退费，累计{len(all_items)}条")

            # 如果本页返回数量少于page_size，说明已到最后一页
            if len(items) < 15:
                break

            page += 1
            if page > 50:  # 安全上限
                break

        logger.info(f"共找到 {len(all_items)} 个【豌豆】退费申请")

        if not all_items:
            logger.info("没有符合条件的数据")
            return

        # 提取每个详情
        for i, item in enumerate(all_items, 1):
            fd_id = item.get('fdId', '')
            subject = item.get('docSubject', '')
            logger.info(f"[{i}/{len(all_items)}] 提取: {subject}")

            html_content = self.lanling.get_detail_page(fd_id)
            data = self.lanling.extract_detail_from_html(html_content, fd_id)
            data['申请单编号'] = item.get('fdNumber', '') or data.get('申请单编号', '')

            self.excel.append(data)
            time.sleep(0.3)  # 避免请求过快

        logger.info(f"\n{'='*60}")
        logger.info(f"提取完成: 共 {len(all_items)} 条数据")
        logger.info(f"{'='*60}\n")


# ============ 入口 ============
def main():
    parser = argparse.ArgumentParser(description="蓝凌OA退费审批自动化工具")
    parser.add_argument("--daily", action="store_true", help="执行每日审批任务")
    parser.add_argument("--extract", action="store_true", help="提取所有历史数据")
    parser.add_argument("--days", type=int, default=30, help="提取最近N天的数据")
    args = parser.parse_args()

    app = RefundApprovalApp()

    if args.daily:
        app.run_daily()
    elif args.extract:
        app.run_extract_all(days=args.days)
    else:
        app.run_daily()


if __name__ == "__main__":
    main()
