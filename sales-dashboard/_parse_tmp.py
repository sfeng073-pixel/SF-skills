
import openpyxl
import json
import sys

def parse_number(val):
    if val is None: return 0
    try: return float(val)
    except: return 0

def parse_percent(val):
    if val is None: return 0
    try:
        v = float(val)
        return v if v <= 1 else v / 100
    except: return 0

def parse_percent_raw(val):
    if val is None: return 0
    try: return float(val)
    except: return 0

def parse_string(val):
    if val is None: return ''
    return str(val).strip()

def find_header_col(ws, field_conf):
    """根据关键词匹配表头列"""
    keywords = field_conf.get('header_keywords', [])
    exclude = field_conf.get('exclude', [])
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col).value or '').strip()
        if all(kw in header for kw in keywords):
            if not any(ek in header for ek in exclude):
                return col
    return None

def parse_summary(ws, config):
    """解析汇总行"""
    result = {}
    match_row_text = config.get('match_row', '')
    
    for row_idx in range(1, min(ws.max_row + 1, 20)):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        row_text = ' '.join([str(c or '') for c in cells])
        
        if match_row_text and match_row_text not in row_text:
            continue
            
        for field_key, field_conf in config.get('fields', {}).items():
            col = find_header_col(ws, field_conf)
            if col:
                val = ws.cell(row=row_idx, column=col).value
                ftype = field_conf.get('type', 'number')
                if ftype == 'percent':
                    result[field_key] = parse_percent(val)
                elif ftype == 'percent_raw':
                    result[field_key] = parse_percent_raw(val)
                else:
                    result[field_key] = parse_number(val)
        
        if match_row_text and match_row_text in row_text:
            break
    
    return result

def parse_extra_metrics(ws, config):
    """解析额外指标（CC跟进等）"""
    result = {}
    match_row = config.get('match_row', 'first_data_row')
    start_row = 2 if match_row == 'first_data_row' else 1
    
    for row_idx in range(start_row, min(start_row + 5, ws.max_row + 1)):
        for field_key, field_conf in config.get('fields', {}).items():
            col = find_header_col(ws, field_conf)
            if col:
                val = ws.cell(row=row_idx, column=col).value
                ftype = field_conf.get('type', 'number')
                if ftype == 'percent_raw':
                    result[field_key] = parse_percent_raw(val)
                else:
                    result[field_key] = parse_number(val)
        if result:
            break
    
    return result

def parse_ranking(ws, config):
    """解析排名列表"""
    items = []
    exclude_keywords = config.get('exclude_rows', [])
    sort_by = config.get('sort_by', 'mtd_rate')
    sort_order = config.get('sort_order', 'desc')
    fields = config.get('fields', {})
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        has_data = False
        
        for field_key, field_conf in fields.items():
            col = find_header_col(ws, field_conf)
            if col:
                val = ws.cell(row=row_idx, column=col).value
                ftype = field_conf.get('type', 'number')
                if ftype == 'percent':
                    row_data[field_key] = parse_percent(val)
                elif ftype == 'string':
                    row_data[field_key] = parse_string(val)
                elif ftype == 'auto_number':
                    row_data[field_key] = int(val) if val and isinstance(val, (int, float)) else 0
                else:
                    row_data[field_key] = parse_number(val)
                if row_data[field_key]:
                    has_data = True
        
        # 检查是否需要排除
        name = row_data.get('name', '')
        if not has_data or not name:
            continue
        if any(kw in name for kw in exclude_keywords):
            continue
        
        items.append(row_data)
    
    # 排序
    reverse = sort_order == 'desc'
    items.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
    
    # 重新编号
    for i, item in enumerate(items):
        item['rank'] = i + 1
    
    return items

def parse_call_data(ws, config):
    """解析通时通次数据"""
    items = []
    sort_by = config.get('sort_by', 'call_time')
    sort_order = config.get('sort_order', 'desc')
    fields = config.get('fields', {})
    
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        has_data = False
        
        for field_key, field_conf in fields.items():
            col = find_header_col(ws, field_conf)
            if col:
                val = ws.cell(row=row_idx, column=col).value
                ftype = field_conf.get('type', 'number')
                if ftype == 'string':
                    row_data[field_key] = parse_string(val)
                else:
                    row_data[field_key] = parse_number(val)
                if row_data[field_key]:
                    has_data = True
        
        name = row_data.get('name', '')
        if not has_data or not name:
            continue
        
        items.append(row_data)
    
    reverse = sort_order == 'desc'
    items.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse)
    
    return items

# === 主流程 ===
files = {"ranking":"/sessions/69fc6f10c9b3ac3b7bf544d9/workspace/益智台湾CC业绩排名.xlsx","followup":"/sessions/69fc6f10c9b3ac3b7bf544d9/workspace/益智CC跟进_台湾.xlsx","call":"/sessions/69fc6f10c9b3ac3b7bf544d9/workspace/通时通次明细.xlsx"}

# 解析汇总指标
summary = {}
if 'ranking' in files:
    print(f'解析汇总指标: {files["ranking"]}')
    wb = openpyxl.load_workbook(files['ranking'], data_only=True)
    summary = parse_summary(wb.active, {"from_report":"ranking","match_row":"台湾益智销售区","fields":{"total_gmv_actual":{"header_keywords":["GMV","实际"],"exclude":["目标"],"type":"number"},"total_gmv_target":{"header_keywords":["GMV","目标"],"exclude":["实际"],"type":"number"},"mtd_rate":{"header_keywords":["MTD","达成率"],"exclude":["单量"],"type":"percent"},"order_actual":{"header_keywords":["单量","实际","加权"],"exclude":["目标","达成率"],"type":"number"},"order_target":{"header_keywords":["单量","目标"],"exclude":["实际","达成率"],"type":"number"},"order_rate":{"header_keywords":["单量","达成率"],"exclude":["GMV"],"type":"percent"}}})
    print(f'  GMV: {summary.get("total_gmv_actual", 0)} / {summary.get("total_gmv_target", 0)}')

# 解析额外指标
if 'followup' in files:
    print(f'解析额外指标: {files["followup"]}')
    wb = openpyxl.load_workbook(files['followup'], data_only=True)
    extra = parse_extra_metrics(wb.active, {"from_report":"followup","match_row":"first_data_row","fields":{"分发转化率":{"header_keywords":["分发转化率"],"exclude":["滚动"],"type":"percent_raw"},"分发滚动转化率":{"header_keywords":["分发滚动转化率"],"exclude":[],"type":"percent_raw"},"滚动ASP":{"header_keywords":["滚动ASP"],"exclude":[],"type":"number"},"产能达成率":{"header_keywords":["产能达成率","产能"],"exclude":[],"type":"percent_raw"}}})
    summary.update(extra)
    print(f'  指标: {len(extra)} 项')

# 解析排名
ranking = []
if 'ranking' in files:
    wb = openpyxl.load_workbook(files['ranking'], data_only=True)
    ranking = parse_ranking(wb.active, {"from_report":"ranking","exclude_rows":["汇总","合计","台湾益智销售区"],"sort_by":"mtd_rate","sort_order":"desc","fields":{"rank":{"header":"排名","type":"auto_number"},"name":{"header_keywords":["姓名","销售"],"type":"string"},"team":{"header_keywords":["团队","组"],"type":"string"},"gmv_actual":{"header_keywords":["GMV","实际"],"exclude":["目标"],"type":"number"},"gmv_target":{"header_keywords":["GMV","目标"],"exclude":["实际"],"type":"number"},"mtd_rate":{"header_keywords":["MTD","达成率"],"exclude":["单量"],"type":"percent"},"order_actual":{"header_keywords":["单量","实际","加权"],"exclude":["目标","达成率"],"type":"number"},"order_target":{"header_keywords":["单量","目标"],"exclude":["实际","达成率"],"type":"number"},"order_rate":{"header_keywords":["单量","达成率"],"exclude":["GMV"],"type":"percent"}}})
    print(f'  排名: {len(ranking)} 人')

# 解析通时通次
call_data = []
if 'call' in files:
    print(f'解析通时通次: {files["call"]}')
    wb = openpyxl.load_workbook(files['call'], data_only=True)
    call_data = parse_call_data(wb.active, {"from_report":"call","sort_by":"call_time","sort_order":"desc","fields":{"name":{"header_keywords":["姓名","名字"],"type":"string"},"call_time":{"header_keywords":["通时"],"type":"number"},"call_count":{"header_keywords":["通次"],"type":"number"}}})
    print(f'  人数: {len(call_data)} 人')

# 保留现有call_data（如果没有新的）
if not call_data:
    try:
        with open('/sessions/6a058fbf579eaa7d4b445082/workspace/sales-dashboard-skill/sales-data.json', 'r', encoding='utf-8') as f:
            existing = json.load(f)
            call_data = existing.get('call_data', [])
        print(f'  保留现有通时通次: {len(call_data)} 人')
    except:
        pass

result = {
    'summary': summary,
    'sales_ranking': ranking,
    'call_data': call_data
}

with open('/sessions/6a058fbf579eaa7d4b445082/workspace/sales-dashboard-skill/sales-data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f'\n✅ 数据已保存到: /sessions/6a058fbf579eaa7d4b445082/workspace/sales-dashboard-skill/sales-data.json')
